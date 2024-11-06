import cv2
import numpy as np
import os
from datetime import datetime
import openpyxl
import re

class ArUcoDetector:
    def __init__(self):
        self.capture_state = None
        self.before_frames = []
        self.after_frames = []
        self.data = []
        self.dictionary = cv2.aruco.getPredefinedDictionary(cv2.aruco.DICT_6X6_250)
        self.parameters = cv2.aruco.DetectorParameters()
        self.target_size = (400, 300)  # Adjust as needed
    # Function to threshold the region of interest
    def threshold_roi(self,image, gap_size):
        # Find connected components in the image
        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(image, connectivity=8)
        
        # Iterate through the connected components
        for label in range(1, num_labels):
            # Get the pixel count of the connected component
            pixel_count = stats[label, cv2.CC_STAT_AREA]
            
            # If the pixel count is less than the gap_size, fill it with black
            if pixel_count < gap_size:
                image[labels == label] = 0
        
        return image
    def detect_markers(self, frame):
        markerCorners, markerIds, _ = cv2.aruco.detectMarkers(frame, self.dictionary, parameters=self.parameters)

        black_pixels = 0
        roi = None
        sorted_corners = []  # Initialize with an empty list
        sorted_ids = []  # Initialize with an empty list
        sorted_indices = []  # Initialize with an empty list
        if markerIds is not None and len(markerIds) >= 4:
            sorted_indices = np.argsort(markerIds.flatten())
            sorted_corners = np.array(markerCorners)[sorted_indices]
            sorted_ids = np.array(markerIds.flatten(), dtype=np.int32)[sorted_indices]

            roi_corners = np.vstack(sorted_corners[:4])[:, 0, :]
            roi_corners = np.array(roi_corners, dtype=np.float32)
            dst_corners = np.array([[0, 0], [self.target_size[0], 0], [self.target_size[0], self.target_size[1]], [0, self.target_size[1]]], dtype=np.float32)
            M = cv2.getPerspectiveTransform(roi_corners, dst_corners)
            roi = cv2.warpPerspective(frame, M, self.target_size)

            gray_image = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, binary_image = cv2.threshold(gray_image, 127, 255, cv2.THRESH_BINARY)
            binary_image[np.all(roi == [0, 0, 255], axis=2)] = 255
            # Fill in ROI black and white image with white pixels less than 300
            binary_image=self.threshold_roi(binary_image,300)
            black_pixels = np.count_nonzero(binary_image == 0)
            # inverted_frame = cv2.bitwise_not(binary_image)
            cv2.imshow("binary_image", binary_image)
            cv2.imshow("ROI", roi)
            cv2.putText(
                frame,
                f"Pixels: {black_pixels}",
                (470, 50),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.5,
                (0, 255, 255),
                2,
            )

            cv2.polylines(frame, [np.int32(roi_corners)], isClosed=True, color=(0, 255, 0), thickness=2)

        if markerIds is not None and len(markerIds) > 0 and len(markerIds) < 4:
            for corners in markerCorners:
                cv2.polylines(frame, [np.int32(corners)], isClosed=True, color=(0, 0, 255), thickness=2)
            cv2.putText(
                frame,
                "Marker missing!",
                (20, 50),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                (0, 0, 255),
                2,
            )

        frame = cv2.aruco.drawDetectedMarkers(frame, np.array(sorted_corners), np.array(sorted_ids))
        return frame, black_pixels, roi, sorted_ids




# ...
    def capture_frames(self, image_folder, save_folder, capture_state):
        if capture_state == 'start':
            self.capture_state = 'before'
            self.before_frames = []
            self.before_timestamps = []
        elif capture_state == 'end':
            self.capture_state = 'after'
            self.after_frames = []
            self.after_timestamps = []

        image_files = sorted(os.listdir(image_folder))
        black_pixels_before = None
        black_pixels_after = None
        roi_before = None
        roi_after = None

        for file in image_files:
            if file.startswith('before_ROI_') and self.capture_state == 'before':
                frame = cv2.imread(os.path.join(image_folder, file))
                frame, black_pixels, roi, markerIds = self.detect_markers(frame)

                if markerIds is not None and len(markerIds) >= 4 and black_pixels is not None:
                    self.before_frames.append(frame)
                    black_pixels_before = black_pixels
                    roi_before = roi

                    # Extract timestamp and angle from the file name
                    file_match = re.search(r'before_ROI_(\d+_\d+)_angle(\d+)', file)
                    if file_match:
                        timestamp = file_match.group(1)
                        angle = int(file_match.group(2))
                    else:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        angle = 0

                    before_filename = f"before_{timestamp}_angle{angle}.jpg"
                    roi_before_filename = f"before_ROI_{timestamp}_angle{angle}.jpg"
                    # cv2.imwrite(os.path.join(save_folder, before_filename), self.before_frames[-1])
                    # cv2.imwrite(os.path.join(save_folder, roi_before_filename), roi_before)
                    self.data.append([timestamp, before_filename, roi_before_filename, angle, black_pixels_before])
                    self.before_timestamps.append(timestamp)

                cv2.imshow("Frame", frame)

            elif file.startswith('after_ROI_') and self.capture_state == 'after':
                frame = cv2.imread(os.path.join(image_folder, file))
                frame, black_pixels, roi, markerIds = self.detect_markers(frame)

                if markerIds is not None and len(markerIds) >= 4 and black_pixels is not None:
                    self.after_frames.append(frame)
                    black_pixels_after = black_pixels
                    roi_after = roi

                    # Extract timestamp and angle from the file name
                    file_match = re.search(r'after_ROI_(\d+_\d+)_angle(\d+)', file)
                    if file_match:
                        timestamp = file_match.group(1)
                        angle = int(file_match.group(2))
                    else:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        angle = 0

                    after_filename = f"after_{timestamp}_angle{angle}.jpg"
                    roi_after_filename = f"after_ROI_{timestamp}_angle{angle}.jpg"
                    # cv2.imwrite(os.path.join(save_folder, after_filename), self.after_frames[-1])
                    # cv2.imwrite(os.path.join(save_folder, roi_after_filename), roi_after)

                    # Find the matching "before" frame based on timestamp
                    index = self.before_timestamps.index(timestamp)
                    difference = abs(black_pixels_after - self.data[index][4])
                    self.data[index].extend([after_filename, roi_after_filename, angle, black_pixels_after, difference])

                cv2.imshow("Frame", frame)

            if cv2.waitKey(20) & 0xFF == ord('q'):
                print(self.data)
                break

        # cap.release()
        cv2.destroyAllWindows()


    def save_data_to_excel(self, excel_file):
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            headers = ['Timestamp', 'Before Filename', 'Before ROI Filename', 'Angle', 'Before Black Pixels', 'After Filename', 'After ROI Filename', 'Angle', 'After Black Pixels', 'Difference']
            sheet.append(headers)

        for row in self.data:
            timestamp = row[0]
            angle = row[3]
            black_pixels_after = row[-2]
            difference = abs(black_pixels_after - row[4])

            # Find the matching row in the Excel file based on timestamp and angle
            for excel_row in sheet.iter_rows(min_row=2, values_only=True):
                if excel_row[0] == timestamp and excel_row[3] == angle:
                    # Update the pixel count and difference
                    excel_row[8] = black_pixels_after
                    excel_row[9] = difference
                    break

        wb.save(excel_file)




# Usage example:
folder = "35"
image_folder = r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\frames__{}".format(folder)
save_folder = r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\frames__{}".format(folder)
excel_file = os.path.join(save_folder, "frame_data_35.xlsx")

os.makedirs(save_folder, exist_ok=True)

aruco_detector = ArUcoDetector()
aruco_detector.capture_frames(image_folder, save_folder, 'start')  # Capture "before" frames with angle 25
aruco_detector.capture_frames(image_folder, save_folder, 'end')  # Capture "after" frames with angle 25
aruco_detector.save_data_to_excel(excel_file)  # Save data to Excel file
