import cv2
import numpy as np
import os
from datetime import datetime
import openpyxl
import re
from openpyxl import load_workbook
class ImageComparator:
    def __init__(self):
        self.data = []

    def compare_images(self, before_image, after_image,angle):
        before_roi = cv2.imread(before_image)
        after_roi = cv2.imread(after_image)

        before_gray = cv2.cvtColor(before_roi, cv2.COLOR_BGR2GRAY)
        after_gray = cv2.cvtColor(after_roi, cv2.COLOR_BGR2GRAY)

        _, before_binary = cv2.threshold(before_gray, 127, 255, cv2.THRESH_BINARY)
        _, after_binary = cv2.threshold(after_gray, 127, 255, cv2.THRESH_BINARY)

        before_binary = self.threshold_roi(before_binary, 300)
        after_binary = self.threshold_roi(after_binary, 300)

        before_black_pixels = np.count_nonzero(before_binary == 0)
        after_black_pixels = np.count_nonzero(after_binary == 0)
        pixel_difference = abs(after_black_pixels - before_black_pixels)

        inverted_before = cv2.bitwise_not(before_binary)
        inverted_after = cv2.bitwise_not(after_binary)

        return before_black_pixels, after_black_pixels, pixel_difference, inverted_before, inverted_after

    def threshold_roi(self, image, gap_size):
        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(image, connectivity=8)

        for label in range(1, num_labels):
            pixel_count = stats[label, cv2.CC_STAT_AREA]
            if pixel_count < gap_size:
                image[labels == label] = 0

        return image


    def extract_angle_timestamp(self,image_filename):
        print(image_filename)
        match = re.search(r'(before|after)_ROI_(\d+)_(\w+)_angle(\d+)\.jpg', image_filename)
        if match:
            timestamp = match.group(2)
            angle = match.group(4)
            return timestamp, angle
        else:
            return None, None

    def capture_images(self, image_folder, save_folder):
        image_files = sorted(os.listdir(image_folder))
        before_images = []
        after_images = []

        for file in image_files:
            print(file)
            if file.startswith('before_ROI_'):
                before_images.append(os.path.join(image_folder, file))
            elif file.startswith('after_ROI_'):
                after_images.append(os.path.join(image_folder, file))

        os.makedirs(save_folder, exist_ok=True)

        for before_image, after_image in zip(before_images, after_images):
            before_filename = os.path.basename(before_image)
            after_filename = os.path.basename(after_image)

            before_timestamp, before_angle = self.extract_angle_timestamp(before_filename)
            after_timestamp, after_angle = self.extract_angle_timestamp(after_filename)
            print(before_timestamp, before_angle)
            if before_timestamp and before_angle:
                try:
                    print("in")
                    before_black_pixels, after_black_pixels, pixel_difference, inverted_before, inverted_after = self.compare_images(before_image, after_image,before_angle)
                    self.data.append([before_timestamp, before_filename, after_filename, before_angle, before_black_pixels, after_black_pixels, pixel_difference])

                    cv2.putText(inverted_before, f"Pixels: {before_black_pixels}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                    cv2.imshow("Inverted Before ROI", inverted_before)

                    cv2.putText(inverted_after, f"Pixels: {after_black_pixels}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                    cv2.imshow("Inverted After ROI", inverted_after)

                    save_folder_path = os.path.join(save_folder, f"binary_{folder}")
                    os.makedirs(save_folder_path, exist_ok=True)

                    before_inverted_filename = f"binary_inverted_before_{before_timestamp}_angle{before_angle}.png"
                    after_inverted_filename = f"binary_inverted_after_{after_timestamp}_angle{after_angle}.png"

                    cv2.imwrite(os.path.join(save_folder_path, before_inverted_filename), inverted_before)
                    cv2.imwrite(os.path.join(save_folder_path, after_inverted_filename), inverted_after)

                    while True:
                        key = cv2.waitKey(1) & 0xFF
                        if key == ord('q'):
                            break

                except Exception as e:
                    print(f"Error processing images: {e}")

        cv2.destroyAllWindows()

    

    def save_data_to_excel(self,excel_file, folder):
            source_file = excel_file
            new_file = f"pixel_count_{folder}.xlsx"
            data=self.data
            # Load the existing workbook
            source_wb = load_workbook(source_file)

            # Create a new workbook as a copy of the existing one
            new_wb = openpyxl.Workbook()
            new_wb.save(new_file)

            # Load the new workbook
            new_wb = load_workbook(new_file)

            # Get the active sheet from both workbooks
            source_sheet = source_wb.active
            new_sheet = new_wb.active

            # Copy the header row from the source sheet to the new sheet
            for col_num, cell in enumerate(source_sheet[1], start=1):
                new_sheet.cell(row=1, column=col_num).value = cell.value

            # Copy the data rows and update the pixel counts
            for row_num in range(2, source_sheet.max_row + 1):
                for col_num, cell in enumerate(source_sheet[row_num], start=1):
                    new_sheet.cell(row=row_num, column=col_num).value = cell.value

                # Update the pixel count values (column indexes may vary, adjust accordingly)
                before_pixel_count = data[row_num - 2][4]  # Assuming pixel count is at index 4
                after_pixel_count = data[row_num - 2][5]  # Assuming pixel count is at index 5
                pixel_difference = data[row_num - 2][6]   # Assuming pixel difference is at index 6

                new_sheet.cell(row=row_num, column=5).value = before_pixel_count
                new_sheet.cell(row=row_num, column=6).value = after_pixel_count
                new_sheet.cell(row=row_num, column=7).value = pixel_difference

            # Save the new workbook
            new_wb.save(new_file)


# Usage example:
folder = "30"
image_folder = r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\frames__{}".format(folder)
save_folder = r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\binary{}_angle".format(folder)
excel_file = os.path.join(image_folder, "frame_data_{}.xlsx".format(folder))

image_comparator = ImageComparator()
image_comparator.capture_images(image_folder, save_folder)  # Compare images and display results
image_comparator.save_data_to_excel(excel_file,folder)  # Save data to Excel file