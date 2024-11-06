import cv2
import numpy as np
import os
import re
import openpyxl
import shutil
# Function to extract angle from image filename
def extract_angle_timestamp(image_filename):
    match = re.search(r'(before|after)_ROI_(\d+)_(\w+)_angle(\d+)\.jpg', image_filename)
    if match:
        timestamp = match.group(2)
        angle = match.group(4)
        return timestamp, angle
    else:
        return None, None

# Function to threshold the region of interest
def threshold_roi(image, gap_size):
    # Find connected components in the image
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(image, connectivity=8)

    # Iterate through the connected components
    for label in range(1, num_labels):
        # Get the pixel count of the connected component
        pixel_count = stats[label, cv2.CC_STAT_AREA]

        # If the pixel count is less than the gap_size, fill it with black
        if pixel_count < gap_size:
            image[labels == label] = 0

    return image

# Interactive console for threshold input
def interactive_console(directory,folder):
    # Get the list of image files in the directory
    files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]

    # Sort the files in ascending order based on the angle value in the filename
    files.sort(key=lambda x: int(extract_angle_timestamp(x)[1]) if extract_angle_timestamp(x)[1] else float('inf'))

    # Create a dictionary to store the pairs of before and after images
    image_pairs = {}
    save_pairs={}
    # Populate the dictionary with image pairs
    for filename in files:
        _, angle = extract_angle_timestamp(filename)
        if angle:
            if angle not in image_pairs:
                image_pairs[angle] = {'before': None, 'after': None}
                save_pairs[angle] = {'before': None, 'after': None}
            if 'before' in filename:
                image_pairs[angle]['before'] = cv2.imread(os.path.join(directory, filename), cv2.IMREAD_GRAYSCALE)
            elif 'after' in filename:
                image_pairs[angle]['after'] = cv2.imread(os.path.join(directory, filename), cv2.IMREAD_GRAYSCALE)

    # Iterate through the image pairs
    while image_pairs:
        # Get the current image pair
        angle = next(iter(image_pairs.keys()))
        print("Test angle: " + angle)
        pair = image_pairs[angle]

        # Convert the images to binary
        _, before_binary = cv2.threshold(pair['before'], 127, 255, cv2.THRESH_BINARY)
        _, after_binary = cv2.threshold(pair['after'], 127, 255, cv2.THRESH_BINARY)

        while True:
            # Input the threshold value
            threshold = int(input("Enter the fill size: "))

            # Threshold the ROIs
            before_thresholded = threshold_roi(before_binary.copy(), threshold)
            after_thresholded = threshold_roi(after_binary.copy(), threshold)
            # Display the current image pair and pixel counts
            before_pixel_count = np.count_nonzero(before_thresholded == 0)
            after_pixel_count = np.count_nonzero(after_thresholded == 0)
            print(f"Before {angle} (Pixel = {before_pixel_count})")
            print(f"After {angle} (Pixel = {after_pixel_count})")
            # Display the updated images
            cv2.imshow(f"Before {angle} (Pixel = {before_pixel_count}) (fill_size = {threshold})", before_thresholded)
            cv2.imshow(f"After {angle} (Pixel = {after_pixel_count}) (fill_size = {threshold})", after_thresholded)
            # gap_fill_directory = os.path.join(directory, "gap_{}".format(folder))
            # os.makedirs(gap_fill_directory, exist_ok=True)
            # save_pairs[angle]['before'] =before_thresholded
            # save_pairs[angle]['after'] =after_thresholded
            # cv2.imwrite(os.path.join(gap_fill_directory, f"before_{angle}.jpg"), before_thresholded)
            # cv2.imwrite(os.path.join(gap_fill_directory, f"after_{angle}.jpg"), after_thresholded)

            # # Remove the current image pair from the dictionary
            # del image_pairs[angle]
            # break
            # Wait for key press
            key = cv2.waitKey(0)
            if key == ord('s'):
                gap_fill_directory = os.path.join(directory, "gap_{}".format(folder))
                os.makedirs(gap_fill_directory, exist_ok=True)
                save_pairs[angle]['before'] =before_thresholded
                save_pairs[angle]['after'] =after_thresholded
                cv2.imwrite(os.path.join(gap_fill_directory, f"before_{angle}.jpg"), before_thresholded)
                cv2.imwrite(os.path.join(gap_fill_directory, f"after_{angle}.jpg"), after_thresholded)

                # Remove the current image pair from the dictionary
                del image_pairs[angle]
                break
            elif key == ord('r'):
                # Reset the threshold value and ask for a new input
                continue

        cv2.destroyAllWindows()

    # Save data to Excel
    save_data_to_excel(directory, save_pairs, angle,folder)

def save_data_to_excel(directory, save_pairs, angle,folder):
    # # excel_file = os.path.join(directory, "frame_data_{}_.xlsx".format(folder))

    # # Load the existing workbook
    # source_wb = openpyxl.load_workbook(excel_file)

    # # Get the active sheet from the workbook
    # source_sheet = source_wb.active

    # # Copy the header row from the source sheet to a new sheet
    new_sheet = openpyxl.Workbook().active
    headers = [ 'Angle', 'Before Black Pixels','After Black Pixels', 'Difference']
    new_sheet.append(headers)
    # for col_num, cell in enumerate(source_sheet[1], start=1):
    #     new_sheet.cell(row=1, column=col_num).value = cell.value

    # Copy the data rows and update the pixel counts
    row_num = 2
    for angle, pair in list(save_pairs.items())[::-1]:
        before_pixel_count = np.count_nonzero(save_pairs[angle]['before'] == 0)
        after_pixel_count = np.count_nonzero(save_pairs[angle]['after'] == 0)

        new_sheet.cell(row=row_num, column=1).value = angle
        new_sheet.cell(row=row_num, column=2).value = before_pixel_count
        new_sheet.cell(row=row_num, column=3).value = after_pixel_count
        new_sheet.cell(row=row_num, column=4).value = before_pixel_count-after_pixel_count

        row_num += 1

    # Save the new workbook
    new_file = os.path.join(directory, "pixel_count_{}.xlsx".format(folder))
    new_sheet.parent.save(new_file)

# Usage example:
folder = "30"
image_folder = r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\frames__{}".format(folder)#r"C:\Users\81809\Desktop\Three_DoF_Robotarm-main\Three_DoF_Robotarm\utils\fig\frames__{}".format(folder)
interactive_console(image_folder,folder)
